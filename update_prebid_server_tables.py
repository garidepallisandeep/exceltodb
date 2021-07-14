import sys, os
import json
import psycopg2
import openpyxl


def createconnection(database, host, user, password, port):
    con = psycopg2.connect(database=database, user=user, password=password, host=host, port=port)
    print("Database connection opened successfully")
    cur = con.cursor()
    return con, cur


def runquery(con, cur, query):
    cur.execute(query)
    con.commit()


def createTable(con, cur, table, columns):
    query1 = []
    for k in columns:
        if k == 'uuid':
            query1.append('{} TEXT PRIMARY KEY'.format(k))
        elif k == 'config':
            query1.append('{} JSONB'.format(k))
        else:
            query1.append('{} TEXT'.format(k))
    finalquery = 'CREATE TABLE IF NOT EXISTS {} ( '.format(table) + ','.join(query1) + ')'
    print('Running Query: {}'.format(finalquery))
    try:
        cur.execute(finalquery)
        con.commit()
    except psycopg2.errors.DuplicateTable:
        print('Issue creating Table : {}'.format(table))
        sys.exit(1)


def insertdata(conn, cursor, table, label, columns, final_json, auction_id):
    basequery = 'INSERT INTO {} '.format(table)
    columnquery = ','.join(list(columns))
    valuequery = []
    try:
        for k in columns:
            if k == 'uuid':
                valuequery.append('\'{}\''.format(auction_id))
            elif k == 'label':
                valuequery.append('\'{}\''.format(label))
            else:
                valuequery.append('\'{}\''.format(json.dumps(final_json)))
        finalquery = '''{} ({}) VALUES ({})'''.format(basequery, columnquery, ','.join(valuequery))
        print(finalquery)
        cursor.execute(finalquery)
        conn.commit()

    except psycopg2.errors.UniqueViolation:
        print('Entry {} already Exists'.format(valuequery[0]))

    except TypeError:
        print(f"JSON file issues")

    except KeyError:
        print(f"JSON file issues")


def selectdata(conn, cursor, table, auction_id):
    finalquery = "select * from {} where uuid = '{}'".format(table, auction_id)
    print('Running {}'.format(finalquery))
    try:
        cursor.execute(finalquery)
        data = cursor.fetchall()
        conn.commit()
        return data
    except psycopg2.errors.InFailedSqlTransaction:
        print("row already exists")


def deletedata(conn, cursor, table, auction_id):
    finalquery = '''DELETE FROM {} where uuid = \'{}\''''.format(table, auction_id)
    print(finalquery)
    cursor.execute(finalquery)
    conn.commit()


def app_data(sheet_name, product_list, file_name, file, pg_db, pg_host, pg_user, pg_password, pg_port, table, label,
             columns):
    original_filename = file.lower()
    print(original_filename)
    for product_row in range(2, product_list.max_row + 1):
        unit_id = product_list.cell(product_row, 4).value
        if unit_id is None:
            pass
        else:
            size_labels = product_list.cell(product_row, 8).value
            size_labels_data = size_labels.lower()
            auction_id = unit_id.replace('/', '-')[1:]
            print(auction_id)
            print(original_filename)
            if "barron" in original_filename:
                if "barrons" in auction_id:
                    print(file_name)
                    newsconnect_placement_id = int(product_list.cell(product_row, 10).value)
                    magnite_account_id = int(product_list.cell(product_row, 11).value)
                    magnite_site_id = int(product_list.cell(product_row, 12).value)
                    magnite_zone_id = int(product_list.cell(product_row, 13).value)
                    if "mrectmobile" in size_labels_data:
                        banner_data = {"format": [{"w": 300, "h": 250}]}
                    else:
                        banner_data = {"format": [{"w": 320, "h": 50}]}
                    ext_dict = {
                        "newsconnect": {
                            "placementId": newsconnect_placement_id,
                        },
                        "rubicon":
                            {
                                "accountId": magnite_account_id,
                                "siteId": magnite_site_id,
                                "zoneId": magnite_zone_id
                            }
                    }
                    final_json = {
                        "id": auction_id,
                        "banner": banner_data,
                        "ext": ext_dict
                    }
                    print(final_json)
                else :
                    final_json = None
            if "marketwatch" in original_filename:
                if "marketwatch" in auction_id:
                    print(file_name)
                    newsconnect_placement_id = int(product_list.cell(product_row, 10).value)
                    magnite_account_id = int(product_list.cell(product_row, 11).value)
                    magnite_site_id = int(product_list.cell(product_row, 12).value)
                    magnite_zone_id = int(product_list.cell(product_row, 13).value)
                    if "mobile_rr" in size_labels_data:
                        banner_data = {"format": [{"w": 300, "h": 250}]}
                    else:
                        banner_data = {"format": [{"w": 320, "h": 50}]}
                    ext_dict = {
                        "newsconnect": {
                            "placementId": newsconnect_placement_id,
                        },
                        "rubicon":
                            {
                                "accountId": magnite_account_id,
                                "siteId": magnite_site_id,
                                "zoneId": magnite_zone_id
                            }
                    }
                    final_json = {
                        "id": auction_id,
                        "banner": banner_data,
                        "ext": ext_dict
                    }
                    print(final_json)
                if "marketwatch" not in auction_id:
                    final_json = None
            if "nypost" in original_filename:
                print(file_name)
                newsconnect_placement_id = int(product_list.cell(product_row, 12).value)
                appnexus_placement_id = int(product_list.cell(product_row, 10).value)
                magnite_account_id = int(product_list.cell(product_row, 16).value)
                magnite_site_id = int(product_list.cell(product_row, 17).value)
                magnite_zone_id = int(product_list.cell(product_row, 18).value)
                tripplelift_inventory_code = product_list.cell(product_row, 30).value
                triplelift_dict = {
                    "inventoryCode": tripplelift_inventory_code
                }
                ext_dict = {
                    "newsconnect": {
                        "placementId": newsconnect_placement_id,
                    },
                    "appnexus": {
                        "placementId": appnexus_placement_id
                    },
                    "rubicon":
                        {
                            "accountId": magnite_account_id,
                            "siteId": magnite_site_id,
                            "zoneId": magnite_zone_id
                        },
                    "triplelift": triplelift_dict
                }
                if "mobile_big" in size_labels_data:
                    banner_data = {"format": [{"w": 300, "h": 250}]}
                else:
                    banner_data = {"format": [{"w": 320, "h": 50}]}
                final_json = {
                    "id": auction_id,
                    "banner": banner_data,
                    "ext": ext_dict
                }
            if final_json is None:
                print("Not a valid data")
            else:
                conn, cursor = createconnection(pg_db, pg_host, pg_user, pg_password, pg_port)
                results = selectdata(conn, cursor, table, auction_id)
                if results is not None:
                    try:
                        deletedata(conn, cursor, table, auction_id)
                        print(f"{auction_id} row dropped")
                        insertdata(conn, cursor, table, label, columns, final_json, auction_id)
                        conn.close()
                    except psycopg2.errors.InFailedSqlTransaction:
                        print("Insert data successful")
                if results is None:
                    insertdata(conn, cursor, table, label, columns, final_json, auction_id)
                    conn.close()
                    print("Insert data successful")

            # conn, cursor = createconnection(pg_db, pg_host, pg_user, pg_password, pg_port)
            # insertdata(conn, cursor, table, label, columns, final_json, auction_id)
            # # print("Insert data successful")
            # # conn.close()
            # conn, cursor = createconnection(pg_db, pg_host, pg_user, pg_password, pg_port)
            # try:
            #     results = selectdata(conn, cursor, table, auction_id)
            #     if results is not None:
            #         try:
            #             deletedata(conn, cursor, table, auction_id)
            #             print(f"{auction_id} row dropped")
            #             insertdata(conn, cursor, table, label, columns, final_json, auction_id)
            #             conn.close()
            #         except psycopg2.errors.InFailedSqlTransaction:
            #             print("Insert data successful")
            #     if results is None:
            #         print("Insert data successful")
            # except psycopg2.errors.InFailedSqlTransaction:
            #     print("Insert data successful")


def amp_data(sheet_name, product_list, file_name, file, pg_db, pg_host, pg_user, pg_password, pg_port, table, label,
             columns):
    original_filename = file.lower()
    for product_row in range(2, product_list.max_row + 1):
        unit_id = product_list.cell(product_row, 5).value
        if unit_id is None:
            pass
        else:
            size_labels = product_list.cell(product_row, 10).value
            size_labels_data = size_labels.lower()
            auction_id = unit_id.replace('/', '-')[1:]
            if "barron" in original_filename:
                if "barrons" in auction_id:
                    print(file_name)
                    newsconnect_placement_id = int(product_list.cell(product_row, 12).value)
                    magnite_account_id = int(product_list.cell(product_row, 13).value)
                    magnite_site_id = int(product_list.cell(product_row, 14).value)
                    magnite_zone_id = int(product_list.cell(product_row, 15).value)
                    site_dict = {
                        "domain": "barrons.com"
                    }
                    if "mobile_big" in size_labels_data or "mrectmobile" in size_labels_data:
                        imp_id = '{}_imp'.format(auction_id)
                        banner_data = {"format": [{"w": 300, "h": 250}]}
                    else:
                        imp_id = '{}_imp'.format(auction_id)
                        banner_data = {"format": [{"w": 320, "h": 50}]}
                    targeting_dict = {"prebid": {"aliases": {"newsconnect": "appnexus"}, "targeting": {
                        "pricegranularity": {"precision": 2,
                                             "buckets": [{"min": 0.01, "max": 10, "increment": 0.01, "cap": False},
                                                         {"min": 10, "max": 20, "increment": 0.05, "cap": False},
                                                         {"min": 20, "max": 95, "increment": 0.5, "cap": True}]}}}}
                    ext_dict = {
                        "newsconnect": {
                            "placementId": newsconnect_placement_id,
                        },
                        "rubicon":
                            {
                                "accountId": magnite_account_id,
                                "siteId": magnite_site_id,
                                "zoneId": magnite_zone_id
                            }
                    }
                    imp_dict = {
                        "id": imp_id,
                        "banner": banner_data,
                        "ext": ext_dict
                    }
                    final_json = {
                        "id": auction_id,
                        "site": site_dict,
                        "ext": targeting_dict,
                        "imp": [
                            imp_dict
                        ]
                    }
                    print(final_json)
                if "barrons" not in auction_id:
                    final_json = None
            if "marketwatch" in original_filename:
                if "marketwatch" in auction_id:
                    print(file_name)
                    newsconnect_placement_id = int(product_list.cell(product_row, 12).value)
                    magnite_account_id = int(product_list.cell(product_row, 13).value)
                    magnite_site_id = int(product_list.cell(product_row, 14).value)
                    magnite_zone_id = int(product_list.cell(product_row, 15).value)
                    site_dict = {
                        "domain": "marketwatch.com"
                    }
                    if "mobile_big" in size_labels_data or "mobile_rr" in size_labels_data:
                        imp_id = '{}_imp'.format(auction_id)
                        banner_data = {"format": [{"w": 300, "h": 250}]}
                    else:
                        imp_id = '{}_imp'.format(auction_id)
                        banner_data = {"format": [{"w": 320, "h": 50}]}
                    targeting_dict = {"prebid": {"aliases": {"newsconnect": "appnexus"}, "targeting": {
                        "pricegranularity": {"precision": 2,
                                             "buckets": [{"min": 0.01, "max": 10, "increment": 0.01, "cap": False},
                                                         {"min": 10, "max": 20, "increment": 0.05, "cap": False},
                                                         {"min": 20, "max": 95, "increment": 0.5, "cap": True}]}}}}
                    ext_dict = {
                        "newsconnect": {
                            "placementId": newsconnect_placement_id,
                        },
                        "rubicon":
                            {
                                "accountId": magnite_account_id,
                                "siteId": magnite_site_id,
                                "zoneId": magnite_zone_id
                            }
                    }
                    imp_dict = {
                        "id": imp_id,
                        "banner": banner_data,
                        "ext": ext_dict
                    }
                    final_json = {
                        "id": auction_id,
                        "site": site_dict,
                        "ext": targeting_dict,
                        "imp": [
                            imp_dict
                        ]
                    }
                    print(final_json)
                if "marketwatch" not in auction_id:
                    final_json = None
            if "nypost" in original_filename:
                print(file_name)
                domain = product_list.cell(product_row, 1).value
                newsconnect_placement_id = int(product_list.cell(product_row, 14).value)
                appnexus_placement_id = int(product_list.cell(product_row, 12).value)
                rubicon_account_id = int(product_list.cell(product_row, 18).value)
                rubicon_site_id = int(product_list.cell(product_row, 19).value)
                rubicon_zone_id = int(product_list.cell(product_row, 20).value)
                tripplelift_inventory_code = product_list.cell(product_row, 31).value
                site_dict = {
                    "domain": domain
                }
                targeting_dict = {"prebid": {"aliases": {"newsconnect": "appnexus"}, "targeting": {
                    "pricegranularity": {"precision": 2, "ranges": [{"min": 0, "max": 3, "increment": 0.01},
                                                                    {"min": 3, "max": 8, "increment": 0.05},
                                                                    {"min": 8, "max": 50, "increment": 0.5}]}}}}
                triplelift_dict = {
                    "inventoryCode": tripplelift_inventory_code
                }
                ext_dict = {
                    "newsconnect": {
                        "placementId": newsconnect_placement_id,
                    },
                    "appnexus": {
                        "placementId": appnexus_placement_id
                    },
                    "rubicon":
                        {
                            "accountId": rubicon_account_id,
                            "siteId": rubicon_site_id,
                            "zoneId": rubicon_zone_id
                        },
                    "triplelift": triplelift_dict
                }
                if "mobile_big" in size_labels_data or "mobile_rr" in size_labels_data:
                    imp_id = '{}_imp'.format(auction_id)
                    banner_data = {"format": [{"w": 300, "h": 250}]}
                else:
                    imp_id = '{}_imp'.format(auction_id)
                    banner_data = {"format": [{"w": 320, "h": 50}]}
                imp_dict = {
                    "id": imp_id,
                    "banner": banner_data,
                    "ext": ext_dict
                }
                final_json = {
                    "id": auction_id,
                    "site": site_dict,
                    "ext": targeting_dict,
                    "imp": [
                        imp_dict
                    ]
                }
            if final_json is None:
                print("Not a valid data")
            else:
                conn, cursor = createconnection(pg_db, pg_host, pg_user, pg_password, pg_port)
                results = selectdata(conn, cursor, table, auction_id)
                if results is not None:
                    try:
                        deletedata(conn, cursor, table, auction_id)
                        print(f"{auction_id} row dropped")
                        insertdata(conn, cursor, table, label, columns, final_json, auction_id)
                        conn.close()
                    except psycopg2.errors.InFailedSqlTransaction:
                        print("Insert data successful")
                if results is None:
                    insertdata(conn, cursor, table, label, columns, final_json, auction_id)
                    conn.close()
                    print("Insert data successful")
            # conn, cursor = createconnection(pg_db, pg_host, pg_user, pg_password, pg_port)
            # insertdata(conn, cursor, table, label, columns, final_json, auction_id)
            # # print("Insert data successful")
            # # conn.close()
            # try:
            #     results = selectdata(conn, cursor, table, auction_id)
            #     if results is not None:
            #         deletedata(conn, cursor, table, auction_id)
            #         print(f"{auction_id} row dropped")
            #         insertdata(conn, cursor, table, label, columns, final_json, auction_id)
            #         conn.close()
            #     if results is None:
            #         print("Insert data successful")


def main(list_of_files):
    pg_host = os.getenv('POSTGRES_HOST', '')
    pg_db = os.getenv('POSTGRES_DATABASE', '')
    pg_user = os.getenv('POSTGRES_USER', '')
    pg_password = os.getenv('POSTGRES_PASSWORD', '')
    pg_port = os.getenv('POSTGRES_PORT', '')
    for file in list_of_files:
        print(file)
        file_name = "/tmp/{}".format(file)
        print(file_name)
        spread_sheet = openpyxl.load_workbook(filename=file_name, data_only=True)
        sheet_list = spread_sheet.sheetnames
        for sheet in sheet_list:
            print(sheet)
            sheet_name = sheet.lower()
            product_list = spread_sheet[sheet]
            if "app" in sheet_name:
                table = 'stored_imps'
                label = 'imp'
                conn, cursor = createconnection(pg_db, pg_host, pg_user, pg_password, pg_port)
                columns = ['uuid', 'config', 'label']
                createTable(conn, cursor, table, columns)
                app_data(sheet_name, product_list, file_name, file, pg_db, pg_host, pg_user, pg_password, pg_port,
                         table, label, columns)
            if "amp" in sheet_name:
                table = 'stored_requests'
                label = 'request'
                conn, cursor = createconnection(pg_db, pg_host, pg_user, pg_password, pg_port)
                columns = ['uuid', 'config', 'label']
                createTable(conn, cursor, table, columns)
                amp_data(sheet_name, product_list, file_name, file, pg_db, pg_host, pg_user, pg_password, pg_port,
                         table, label, columns)


if __name__ == '__main__':
    main(list_of_files)